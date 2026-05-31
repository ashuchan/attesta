from __future__ import annotations

import io
from typing import Any

from sourceloop.parsing.base import ParseSource, RawRowSet

XLSX_MAGIC = b"PK\x03\x04"
BOM_HEADER_KEYWORDS = {
    "mpn", "part number", "partnumber", "part_number",
    "qty", "quantity", "ref", "designator", "description",
    "manufacturer", "mfr", "brand", "value",
}


class XlsxParser:
    key = "xlsx"
    priority = 10
    enabled = True

    def supports(self, source: ParseSource) -> float:
        if source.content[:4] == XLSX_MAGIC:
            return 0.95
        if source.filename.lower().endswith(".xlsx"):
            return 0.7
        return 0.0

    def parse(self, source: ParseSource) -> RawRowSet:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(source.content), data_only=True, read_only=True)
        best_sheet = self._pick_bom_sheet(wb)
        rows = self._extract_rows(best_sheet)
        wb.close()
        return RawRowSet(
            rows=rows,
            detected_format="xlsx",
            parser_key=self.key,
            parser_confidence=0.9,
        )

    def _pick_bom_sheet(self, wb: Any) -> Any:
        """Pick the sheet most likely to be a BOM based on header matches."""
        best_sheet = None
        best_score = -1
        for name in wb.sheetnames:
            ws = wb[name]
            score = self._score_sheet(ws)
            if score > best_score:
                best_score = score
                best_sheet = ws
        return best_sheet

    def _score_sheet(self, ws: Any) -> int:
        """Score a sheet by counting BOM header keyword matches in first 5 rows."""
        score = 0
        for row in ws.iter_rows(max_row=5, values_only=True):
            for cell in row:
                if cell and str(cell).strip().lower() in BOM_HEADER_KEYWORDS:
                    score += 1
        return score

    def _extract_rows(self, ws: Any) -> list[dict[str, str | None]]:
        rows = []
        headers: list[str] = []
        header_row_found = False
        for row in ws.iter_rows(values_only=True):
            values = [str(c).strip() if c is not None else "" for c in row]
            if not any(values):
                continue
            if not header_row_found:
                # Detect header row: majority of cells match BOM keywords
                matches = sum(1 for v in values if v.lower() in BOM_HEADER_KEYWORDS)
                if matches >= 2 or (len(values) >= 2 and matches >= 1):
                    headers = [v.lower() for v in values]
                    header_row_found = True
                    continue
            if header_row_found and headers:
                row_dict: dict[str, str | None] = {}
                for i, h in enumerate(headers):
                    row_dict[h] = values[i] if i < len(values) else None
                if any(v for v in row_dict.values() if v):
                    rows.append(row_dict)
        return rows
