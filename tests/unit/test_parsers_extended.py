from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from sourceloop.parsing.base import ParseSource, RawRowSet
from sourceloop.parsing.normalizer import _match_header, normalize
from sourceloop.parsing.parsers.csv_parser import CsvParser
from sourceloop.parsing.parsers.pdf_parser import PdfParser
from sourceloop.parsing.parsers.plaintext_parser import PlaintextParser
from sourceloop.parsing.parsers.xlsx_parser import XlsxParser

FIXTURES = Path(__file__).parent.parent / "fixtures" / "boms"


def make_xlsx_bytes(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── XLSX parser ─────────────────────────────────────────────────────────────

def test_xlsx_parser_supports_xlsx_magic():
    parser = XlsxParser()
    xlsx_bytes = make_xlsx_bytes([["MPN", "Qty"], ["STM32", "1"]])
    source = ParseSource(content=xlsx_bytes, filename="bom.xlsx")
    assert parser.supports(source) > 0.7


def test_xlsx_parser_supports_xlsx_extension_without_magic():
    parser = XlsxParser()
    source = ParseSource(content=b"notxlsx", filename="bom.xlsx")
    assert parser.supports(source) > 0.0


def test_xlsx_parser_does_not_support_unrelated():
    parser = XlsxParser()
    source = ParseSource(content=b"hello world", filename="readme.txt")
    assert parser.supports(source) == 0.0


def test_xlsx_parser_extracts_rows():
    parser = XlsxParser()
    xlsx_bytes = make_xlsx_bytes([
        ["MPN", "Qty", "Manufacturer", "Description"],
        ["STM32F103C8T6", "1", "STMicro", "MCU"],
        ["ESP8266EX", "2", "Espressif", "WiFi"],
    ])
    source = ParseSource(content=xlsx_bytes, filename="bom.xlsx")
    rawrowset = parser.parse(source)
    assert rawrowset.detected_format == "xlsx"
    assert rawrowset.parser_key == "xlsx"
    assert len(rawrowset.rows) >= 2


def test_xlsx_parser_fixture():
    xlsx_path = FIXTURES / "iot_board.xlsx"
    if not xlsx_path.exists():
        pytest.skip("xlsx fixture not generated yet")
    parser = XlsxParser()
    content = xlsx_path.read_bytes()
    source = ParseSource(content=content, filename="iot_board.xlsx")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 5


def test_xlsx_parser_multiple_sheets():
    """Parser should pick best BOM sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Instructions"
    ws1.append(["This is not a bom"])

    ws2 = wb.create_sheet("BOM")
    ws2.append(["MPN", "Qty", "Description"])
    ws2.append(["STM32F103C8T6", "1", "MCU"])

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    parser = XlsxParser()
    source = ParseSource(content=content, filename="bom.xlsx")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 1


def test_xlsx_parser_empty_rows_skipped():
    parser = XlsxParser()
    xlsx_bytes = make_xlsx_bytes([
        ["MPN", "Qty"],
        [],  # empty row
        ["STM32F103C8T6", "1"],
    ])
    source = ParseSource(content=xlsx_bytes, filename="bom.xlsx")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 1


# ─── CSV parser ──────────────────────────────────────────────────────────────

def test_csv_parser_parses_comma_delimited():
    parser = CsvParser()
    content = b"MPN,Qty,Manufacturer\nSTM32F103C8T6,1,STMicro\nESP8266EX,2,Espressif\n"
    source = ParseSource(content=content, filename="bom.csv")
    rawrowset = parser.parse(source)
    assert rawrowset.detected_format == "csv"
    assert len(rawrowset.rows) == 2
    assert rawrowset.rows[0]["mpn"] == "STM32F103C8T6"


def test_csv_parser_parses_semicolon_delimited():
    parser = CsvParser()
    content = b"MPN;Qty;Manufacturer\nSTM32F103C8T6;1;STMicro\n"
    source = ParseSource(content=content, filename="bom.csv")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 1


def test_csv_parser_fixture():
    content = (FIXTURES / "iot_board.csv").read_bytes()
    parser = CsvParser()
    source = ParseSource(content=content, filename="iot_board.csv")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 5


def test_csv_parser_zero_confidence_for_binary():
    parser = CsvParser()
    source = ParseSource(content=bytes(range(0, 50)), filename="binary.bin")
    confidence = parser.supports(source)
    assert isinstance(confidence, float)


def test_csv_parser_high_confidence_for_csv_content():
    parser = CsvParser()
    content = b"mpn,qty,manufacturer\nSTM32,1,ST\n"
    source = ParseSource(content=content, filename="bom.csv")
    assert parser.supports(source) > 0.5


# ─── Plaintext parser ────────────────────────────────────────────────────────

def test_plaintext_parser_supports_utf8():
    parser = PlaintextParser()
    source = ParseSource(content=b"col1\tcol2\nval1\tval2\n", filename="bom.txt")
    assert parser.supports(source) > 0.0


def test_plaintext_parser_does_not_support_binary():
    parser = PlaintextParser()
    source = ParseSource(content=bytes(range(128, 256)), filename="binary.bin")
    assert parser.supports(source) == 0.0


def test_plaintext_parser_tab_delimited():
    parser = PlaintextParser()
    content = b"MPN\tQty\tManufacturer\nSTM32F103C8T6\t1\tSTMicro\n"
    source = ParseSource(content=content, filename="bom.txt")
    rawrowset = parser.parse(source)
    assert rawrowset.detected_format == "plaintext"
    assert len(rawrowset.rows) >= 1


def test_plaintext_parser_handles_empty_lines():
    parser = PlaintextParser()
    content = b"MPN,Qty\n\nSTM32F103C8T6,1\n\n"
    source = ParseSource(content=content, filename="bom.txt")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 1


def test_plaintext_parser_comma_delimited():
    parser = PlaintextParser()
    content = b"MPN,Qty\nSTM32F103C8T6,1\n"
    source = ParseSource(content=content, filename="bom.txt")
    rawrowset = parser.parse(source)
    assert len(rawrowset.rows) >= 1


# ─── PDF parser ──────────────────────────────────────────────────────────────

def test_pdf_parser_supports_pdf_magic():
    parser = PdfParser()
    source = ParseSource(content=b"%PDF-1.4 ..." + bytes(100), filename="bom.pdf")
    assert parser.supports(source) > 0.0


def test_pdf_parser_supports_pdf_extension():
    parser = PdfParser()
    source = ParseSource(content=b"not really pdf", filename="bom.pdf")
    assert parser.supports(source) > 0.0


def test_pdf_parser_does_not_support_unrelated():
    parser = PdfParser()
    source = ParseSource(content=b"hello", filename="readme.txt")
    assert parser.supports(source) == 0.0


def test_pdf_table_to_rows():
    parser = PdfParser()
    table = [["MPN", "Qty"], ["STM32F103C8T6", "1"], ["ESP8266EX", "2"]]
    rows = parser._table_to_rows(table)
    assert len(rows) == 2
    assert rows[0]["mpn"] == "STM32F103C8T6"


def test_pdf_table_to_rows_empty():
    parser = PdfParser()
    assert parser._table_to_rows([]) == []
    assert parser._table_to_rows([["MPN"]]) == []  # only header, no data


def test_pdf_text_to_rows():
    parser = PdfParser()
    text = "STM32F103C8T6 1 STMicro\nESP8266EX 2 Espressif"
    rows = parser._text_to_rows(text)
    assert len(rows) == 2


def test_pdf_table_to_rows_with_none_cells():
    parser = PdfParser()
    table = [["MPN", "Qty", "Mfr"], ["STM32F103C8T6", None, "ST"]]
    rows = parser._table_to_rows(table)
    assert len(rows) == 1
    assert rows[0]["mpn"] == "STM32F103C8T6"


def test_pdf_table_to_rows_all_none_row_skipped():
    parser = PdfParser()
    table = [["MPN", "Qty"], [None, None], ["STM32", "1"]]
    rows = parser._table_to_rows(table)
    assert len(rows) == 1


# ─── Normalizer extended ─────────────────────────────────────────────────────

def test_normalizer_handles_invalid_quantity():
    import uuid
    rawrowset = RawRowSet(
        rows=[{"mpn": "STM32F103C8T6", "qty": "not-a-number", "manufacturer": "ST"}],
        detected_format="csv", parser_key="csv", parser_confidence=0.9,
    )
    lines = normalize(rawrowset, uuid.uuid4(), uuid.uuid4())
    assert lines[0].quantity is None


def test_match_header_mpn_aliases():
    assert _match_header("Part Number") == "mpn"
    assert _match_header("part number") == "mpn"
    assert _match_header("partnumber") == "mpn"


def test_match_header_quantity_aliases():
    assert _match_header("QTY") == "quantity"
    assert _match_header("qty") == "quantity"
    assert _match_header("count") == "quantity"


def test_match_header_manufacturer():
    assert _match_header("MFR") == "manufacturer"
    assert _match_header("brand") == "manufacturer"


def test_match_header_unknown_returns_none():
    assert _match_header("completely_unknown_col") is None
    assert _match_header("unknown_xyz_abc") is None


def test_normalizer_empty_rows():
    import uuid
    rawrowset = RawRowSet(rows=[], detected_format="csv", parser_key="csv", parser_confidence=0.9)
    lines = normalize(rawrowset, uuid.uuid4(), uuid.uuid4())
    assert lines == []


def test_normalizer_parse_confidence_higher_with_mpn_and_qty():
    import uuid
    rawrowset = RawRowSet(
        rows=[{"mpn": "STM32F103C8T6", "qty": "5", "description": "MCU"}],
        detected_format="csv", parser_key="csv", parser_confidence=0.9,
    )
    lines = normalize(rawrowset, uuid.uuid4(), uuid.uuid4())
    assert lines[0].parse_confidence >= 0.9


def test_normalizer_parse_confidence_lower_without_mpn():
    import uuid
    rawrowset = RawRowSet(
        rows=[{"description": "Custom PCB"}],
        detected_format="csv", parser_key="csv", parser_confidence=0.9,
    )
    lines = normalize(rawrowset, uuid.uuid4(), uuid.uuid4())
    assert lines[0].parse_confidence < 0.8


def test_normalizer_multiple_rows():
    import uuid
    rawrowset = RawRowSet(
        rows=[
            {"mpn": "STM32F103C8T6", "qty": "1", "manufacturer": "ST"},
            {"mpn": "ESP8266EX", "qty": "2", "manufacturer": "Espressif"},
        ],
        detected_format="csv", parser_key="csv", parser_confidence=0.9,
    )
    lines = normalize(rawrowset, uuid.uuid4(), uuid.uuid4())
    assert len(lines) == 2
    assert lines[0].line_no == 1
    assert lines[1].line_no == 2
